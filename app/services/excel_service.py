from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelService:

    def __init__(self):

        self.output_folder = Path("output")
        self.output_folder.mkdir(exist_ok=True)

        self.file_path = self.output_folder / "PTA_Report.xlsx"

        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "PTA Report"

        self.create_header()

    def create_header(self):

        headers = [

            "File Name",
            "Date",
            "Duration (min)",
            "From (time)",
            "To (time)",
            "Sitting Time (min)",
            "Frame Count",

        ]

        for col, header in enumerate(headers, start=1):

            cell = self.sheet.cell(row=1, column=col)

            cell.value = header
            cell.font = Font(bold=True)

    def frame_to_time(self, frame, fps):

        total_seconds = frame / fps

        hours = int(total_seconds // 3600)

        minutes = int((total_seconds % 3600) // 60)

        seconds = int(total_seconds % 60)

        return f"{hours:02}:{minutes:02}:{seconds:02}"

    def add_video(

        self,

        file_name,

        duration,

        fps,

        frame_count,

        sessions,

    ):

        duration_min = round(duration / 60.0, 1)

        total_sitting = 0

        today = datetime.now().strftime("%Y-%m-%d")

        if len(sessions) == 0:

            row = self.sheet.max_row + 1

            self.sheet.cell(row=row, column=1).value = file_name
            self.sheet.cell(row=row, column=2).value = today
            self.sheet.cell(row=row, column=3).value = duration_min
            self.sheet.cell(row=row, column=4).value = "-"
            self.sheet.cell(row=row, column=5).value = "-"
            self.sheet.cell(row=row, column=6).value = 0
            self.sheet.cell(row=row, column=7).value = frame_count

            return

        first = True

        for session in sessions:

            start_frame = session["start_frame"]

            end_frame = session["end_frame"]

            sitting_frames = end_frame - start_frame + 1

            sitting_minutes = round(

                (sitting_frames / fps) / 60.0,

                1,

            )

            total_sitting += sitting_minutes

            row = self.sheet.max_row + 1

            if first:

                self.sheet.cell(row=row, column=1).value = file_name
                self.sheet.cell(row=row, column=2).value = today
                self.sheet.cell(row=row, column=3).value = duration_min

                first = False

            self.sheet.cell(row=row, column=4).value = self.frame_to_time(
                start_frame,
                fps,
            )

            self.sheet.cell(row=row, column=5).value = self.frame_to_time(
                end_frame,
                fps,
            )

            self.sheet.cell(row=row, column=6).value = sitting_minutes

            self.sheet.cell(row=row, column=7).value = sitting_frames

        row = self.sheet.max_row + 1

        self.sheet.cell(row=row, column=1).value = "TOTAL"

        self.sheet.cell(row=row, column=3).value = duration_min

        self.sheet.cell(row=row, column=6).value = round(
            total_sitting,
            1,
        )

        self.sheet.cell(row=row, column=1).font = Font(bold=True)

        self.sheet.cell(row=row, column=3).font = Font(bold=True)

        self.sheet.cell(row=row, column=6).font = Font(bold=True)

    def save(self):

        self.workbook.save(self.file_path)

    def get_file_path(self):

        return self.file_path